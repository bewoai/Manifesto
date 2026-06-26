from __future__ import annotations

import sqlite3
import pytest
from pathlib import Path
import openpyxl

from app.manifest.importer import sheet_name_to_iso, import_excel_to_sqlite
from app.db import connect

def test_sheet_name_to_iso():
    assert sheet_name_to_iso("22.06.2026") == "2026-06-22"
    assert sheet_name_to_iso("01.01.2020") == "2020-01-01"
    assert sheet_name_to_iso("3.5.25") == "2025-05-03"
    assert sheet_name_to_iso("Sayfa2") is None
    assert sheet_name_to_iso("ULKELER") is None
    assert sheet_name_to_iso("32.13.2026") is None # Invalid date

def test_import_excel_to_sqlite(tmp_path):
    # 1. Create a dummy planning Excel file
    excel_path = tmp_path / "planning_test.xlsx"
    wb = openpyxl.Workbook()
    
    # We need a date sheet name
    ws = wb.active
    ws.title = "22.06.2026"
    
    # Add headers on row 3 (planning sheet uses row 3 for headers, data on row 4+)
    # Column mappings:
    # 1: PAX, 2: UYRUK, 3: M/F, 4: REHBER & MÜŞTERİ İSMİ, 5: İRTİBAT VE ODA NO, 6: OTEL,
    # 7: PICK-UP, 8: REZERVE YAPAN, 9: ACENTE, 10: UÇACAĞI FİRMA, 11: BALON, 12: PİLOT,
    # 13: NOT, 14: ALIŞ ŞÖFÖR, 15: GELECEĞİ YER, 20: PASSAPORT NO
    ws.cell(row=3, column=1, value="PAX")
    ws.cell(row=3, column=2, value="UYRUK")
    ws.cell(row=3, column=3, value="M/F")
    ws.cell(row=3, column=4, value="REHBER & MÜŞTERİ İSMİ")
    ws.cell(row=3, column=6, value="OTEL")
    ws.cell(row=3, column=7, value="PICK-UP")
    ws.cell(row=3, column=9, value="ACENTE")
    ws.cell(row=3, column=11, value="BALON")
    ws.cell(row=3, column=20, value="PASSAPORT NO")
    
    # Row 4: Reservation 1, Pax 2, first passenger (lead)
    ws.cell(row=4, column=1, value="2")
    ws.cell(row=4, column=2, value="TUR")
    ws.cell(row=4, column=3, value="M")
    ws.cell(row=4, column=4, value="AHMET YILMAZ")
    ws.cell(row=4, column=6, value="KAYA HOTEL")
    ws.cell(row=4, column=7, value="05:00")
    ws.cell(row=4, column=9, value="ACME TRAVEL")
    ws.cell(row=4, column=11, value="BZR")
    ws.cell(row=4, column=20, value="TR123456")
    
    # Row 5: Reservation 1, second passenger (PAX empty)
    ws.cell(row=5, column=2, value="TUR")
    ws.cell(row=5, column=3, value="F")
    ws.cell(row=5, column=4, value="AYSE YILMAZ")
    ws.cell(row=5, column=20, value="TR789012")
    
    # Row 6: Reservation 2, Pax 1
    ws.cell(row=6, column=1, value="1")
    ws.cell(row=6, column=2, value="FRA")
    ws.cell(row=6, column=3, value="F")
    ws.cell(row=6, column=4, value="JEANNE DARC")
    ws.cell(row=6, column=6, value="MARS HOTEL")
    ws.cell(row=6, column=7, value="05:15")
    ws.cell(row=6, column=9, value="PARIS TOUR")
    ws.cell(row=6, column=11, value="BYF")
    ws.cell(row=6, column=20, value="FR999888")
    
    wb.save(excel_path)
    wb.close()
    
    # 2. Setup temporary SQLite DB path
    db_path = tmp_path / "test_run.db"
    
    # Initialize DB (which runs schemas)
    from app.db import init_db
    init_db(db_path)
    
    # 3. Perform import
    stats = import_excel_to_sqlite(db_path=db_path, planning_xlsx=excel_path, sheet_name="22.06.2026")
    
    # Assert stats returned
    assert stats["status"] == "success"
    assert stats["passengers_added"] == 3
    assert stats["reservations_created"] == 2
    assert stats["total_rows"] == 3
    
    # 4. Check SQLite DB records
    conn = connect(db_path)
    
    # Check flight
    flights = conn.execute("SELECT * FROM flights").fetchall()
    assert len(flights) == 1
    assert flights[0]["flight_date"] == "2026-06-22"
    flight_id = flights[0]["id"]
    
    # Check reservations
    res = conn.execute("SELECT * FROM reservations WHERE flight_id = ? ORDER BY sort_order", (flight_id,)).fetchall()
    assert len(res) == 2
    assert res[0]["pax"] == 2
    assert res[0]["hotel"] == "KAYA HOTEL"
    assert res[0]["balloon_code"] == "BZR"
    assert res[1]["pax"] == 1
    assert res[1]["hotel"] == "MARS HOTEL"
    assert res[1]["balloon_code"] == "BYF"
    
    # Check passengers
    passengers = conn.execute("SELECT * FROM passengers ORDER BY id").fetchall()
    assert len(passengers) == 3
    assert passengers[0]["full_name"] == "AHMET YILMAZ"
    assert passengers[0]["passport_no"] == "TR123456"
    assert passengers[1]["full_name"] == "AYSE YILMAZ"
    assert passengers[1]["passport_no"] == "TR789012"
    assert passengers[2]["full_name"] == "JEANNE DARC"
    assert passengers[2]["passport_no"] == "FR999888"
    
    # Check reservation_passengers link
    links = conn.execute("SELECT * FROM reservation_passengers WHERE reservation_id = ?", (res[0]["id"],)).fetchall()
    assert len(links) == 2
    assert links[0]["passenger_id"] == passengers[0]["id"]
    assert links[0]["seq"] == 1
    assert links[1]["passenger_id"] == passengers[1]["id"]
    assert links[1]["seq"] == 2
    
    # 5. Test Upsert: Import again with some updates
    # Change first passenger's name and nationality, and add a passenger without passport
    wb2 = openpyxl.load_workbook(excel_path)
    ws2 = wb2["22.06.2026"]
    # Update Ahmet's name spelling and nationality in Excel
    ws2.cell(row=4, column=4, value="AHMET C. YILMAZ")
    ws2.cell(row=4, column=2, value="USA")
    
    # Add new Reservation 3: Pax 1 without passport (empty passport_no)
    ws2.cell(row=7, column=1, value="1")
    ws2.cell(row=7, column=2, value="GER")
    ws2.cell(row=7, column=3, value="M")
    ws2.cell(row=7, column=4, value="JOHN DOE")
    ws2.cell(row=7, column=11, value="BZR")
    ws2.cell(row=7, column=20, value="")
    
    wb2.save(excel_path)
    wb2.close()
    
    # Run import again
    stats2 = import_excel_to_sqlite(db_path=db_path, planning_xlsx=excel_path, sheet_name="22.06.2026")
    
    assert stats2["status"] == "success"
    # Ahmet and Ayse and Jeanne are already in DB. Ahmet gets updated.
    # John Doe gets added.
    assert stats2["passengers_added"] == 1 # John Doe
    assert stats2["passengers_updated"] == 3 # All 3 existing passengers matched and updated
    assert stats2["reservations_created"] == 3
    
    # Verify Ahmet's update in DB
    p_ahmet = conn.execute("SELECT * FROM passengers WHERE passport_no = 'TR123456'").fetchone()
    assert p_ahmet["full_name"] == "AHMET C. YILMAZ"
    assert p_ahmet["nationality"] == "USA"
    
    # Verify John Doe exists
    p_john = conn.execute("SELECT * FROM passengers WHERE full_name = 'JOHN DOE'").fetchone()
    assert p_john is not None
    assert p_john["passport_no"] is None
    
    conn.close()
