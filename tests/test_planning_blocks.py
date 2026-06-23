import openpyxl

from app import config
from app.manifest.planning import PlanningRow, delete_passenger_from_reservation, group_blocks, read_blocks


def _row(row: int, pax: int | None, name: str, passport_no: str) -> PlanningRow:
    return PlanningRow(
        row=row,
        pax=pax,
        nationality="CHN",
        sex="F",
        name=name,
        room="",
        hotel="ARGOS",
        pickup="04:10",
        reserved_by="MAHMUT",
        agency="RANK UNITED",
        company="",
        balloon="BYF",
        pilot="",
        note="",
        driver="",
        coming_place="",
        passport_no=passport_no,
    )


def test_group_blocks_keeps_all_passenger_identities():
    blocks = group_blocks([
        _row(4, 2, "FIRST WOMAN", "P1"),
        _row(5, None, "SECOND WOMAN", "P2"),
    ])

    assert len(blocks) == 1
    assert blocks[0].rows == [4, 5]
    assert [p.name for p in blocks[0].passengers] == ["FIRST WOMAN", "SECOND WOMAN"]
    assert [p.passport_no for p in blocks[0].passengers] == ["P1", "P2"]


def _workbook(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "23.06.2026"
    ws.cell(4, config.COL_PAX).value = 2
    ws.cell(4, config.COL_NAME).value = "FIRST WOMAN"
    ws.cell(4, config.COL_PASSPORT_NO).value = "P1"
    ws.cell(4, config.COL_HOTEL).value = "ARGOS"
    ws.cell(4, config.COL_PICKUP).value = "04:10"
    ws.cell(4, config.COL_AGENCY).value = "ACME"
    ws.cell(4, config.COL_BALLOON).value = "BYF"
    ws.cell(5, config.COL_NAME).value = "SECOND WOMAN"
    ws.cell(5, config.COL_PASSPORT_NO).value = "P2"
    ws.cell(5, config.COL_BALLOON).value = "BYF"
    ws.merge_cells(start_row=4, start_column=config.COL_PAX, end_row=5, end_column=config.COL_PAX)
    ws.merge_cells(start_row=4, start_column=config.COL_HOTEL, end_row=5, end_column=config.COL_HOTEL)
    wb.save(path)
    wb.close()


def test_delete_non_lead_passenger_decrements_pax(tmp_path):
    path = tmp_path / "planning.xlsx"
    _workbook(path)

    delete_passenger_from_reservation(path, "23.06.2026", lead_row=4, rows=[4, 5], target_row=5)
    block = read_blocks(path, "23.06.2026")[0]

    assert block.pax == 1
    assert [p.name for p in block.passengers] == ["FIRST WOMAN"]


def test_delete_lead_passenger_moves_next_identity_to_lead(tmp_path):
    path = tmp_path / "planning.xlsx"
    _workbook(path)

    delete_passenger_from_reservation(path, "23.06.2026", lead_row=4, rows=[4, 5], target_row=4)
    block = read_blocks(path, "23.06.2026")[0]

    assert block.pax == 1
    assert block.hotel == "ARGOS"
    assert block.agency == "ACME"
    assert [p.name for p in block.passengers] == ["SECOND WOMAN"]
    assert [p.passport_no for p in block.passengers] == ["P2"]
