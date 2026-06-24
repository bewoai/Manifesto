import os
import calendar
from pathlib import Path
import openpyxl

from app.flight_plan_generator import generate_monthly_flight_plan, MONTHS_TR
from app import config

def test_monthly_flight_plan_generation():
    # Ayarla
    year = 2026
    month = 7
    tmp_path = Path("tmp")
    tmp_path.mkdir(exist_ok=True)
    save_path = tmp_path / "test_ucus_plani.xlsx"
    if save_path.exists():
        save_path.unlink()
    
    # Oluştur
    generate_monthly_flight_plan(year, month, str(save_path))
    
    # Doğrula
    assert save_path.exists()
    wb = openpyxl.load_workbook(save_path)
    
    # Temmuz ayı 31 gündür
    _, num_days = calendar.monthrange(year, month)
    assert num_days == 31
    assert len(wb.sheetnames) == 31
    
    # İlk sayfa kontrolü
    ws1 = wb["01.07.2026"]
    
    import datetime
    header_found = False
    for r in range(1, 4):
        for c in range(1, 21):
            val = ws1.cell(row=r, column=c).value
            if isinstance(val, datetime.datetime) and val.year == 2026 and val.month == 7 and val.day == 1:
                header_found = True
                break
    assert header_found, "Ay ismi ve yılı başlığa başarıyla yazılmalı"
    
    assert "TEMPLATE" not in wb.sheetnames

def test_leap_year_generation():
    year = 2028
    month = 2
    tmp_path = Path("tmp")
    tmp_path.mkdir(exist_ok=True)
    save_path = tmp_path / "test_leap.xlsx"
    if save_path.exists():
        save_path.unlink()
    
    generate_monthly_flight_plan(year, month, str(save_path))
    
    wb = openpyxl.load_workbook(save_path)
    assert len(wb.sheetnames) == 29
    assert "29.02.2028" in wb.sheetnames

if __name__ == '__main__':
    test_monthly_flight_plan_generation()
    test_leap_year_generation()
    print('All tests passed!')
