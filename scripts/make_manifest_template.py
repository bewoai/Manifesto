"""Mevcut (dolu) bir manifesto dosyasından PII'siz şablon üretir.

Kaynak BZR.xlsx zaten doldurulmuş; bu script MANİFESTO veri satırlarını
temizleyip (header + biçim + ULKELER sayfası + varsa data validation korunur)
data/manifest_template.xlsx olarak yazar. Şablon PII içermez, repo'ya konabilir.

Kullanım:
  python -m scripts.make_manifest_template [--source <xlsx>] [--out <xlsx>]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app import config  # noqa: E402


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", type=Path, default=config.SOURCE_MANIFEST_XLSX)
    ap.add_argument("--out", type=Path, default=config.MANIFEST_TEMPLATE_PATH)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.source)
    ws = wb[config.MANIFEST_SHEET]
    cleared = 0
    for r in range(config.MANIFEST_FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None
                cleared += 1
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    wb.close()
    print(f"Şablon yazıldı -> {args.out} ({cleared} hücre temizlendi, PII kaldırıldı)")


if __name__ == "__main__":
    main()
