import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

src = 'app/templates/monthly_flight_template.xlsx'
wb = openpyxl.load_workbook(src)
ws = wb['TEMPLATE']

# 1. Freeze Panes
ws.freeze_panes = 'A4' # Freezes rows 1, 2, 3

# 2. Data Validation
dv_mf = DataValidation(type='list', formula1='"M,F"', allow_blank=True)
ws.add_data_validation(dv_mf)
dv_mf.add('C4:C500')

dv_balloon = DataValidation(type='list', formula1='"BYF,BTK,BZR,BZV,BYJ"', allow_blank=True)
ws.add_data_validation(dv_balloon)
dv_balloon.add('K4:K500')

dv_driver = DataValidation(type='list', formula1='"1,2,3,4,5"', allow_blank=True)
ws.add_data_validation(dv_driver)
dv_driver.add('N4:N500')

# 3. Conditional formatting
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
rule = FormulaRule(formula=['AND($D4<>"", $T4="")'], stopIfTrue=True, fill=red_fill)
ws.conditional_formatting.add('T4:T500', rule)

wb.save(src)
print('Excel template enhanced!')
