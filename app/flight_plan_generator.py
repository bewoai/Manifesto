from __future__ import annotations

import calendar
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app import config

MONTHS_TR = [
    "", "OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
    "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"
]

def generate_monthly_flight_plan(year: int, month: int, save_path: str) -> None:
    """
    Belirtilen yıl ve ay için şablon Excel dosyasından yeni bir aylık uçuş planı üretir.
    Şablon: app/templates/monthly_flight_template.xlsx
    - İçindeki TEMPLATE sayfasını gün sayısına göre (28, 29, 30, 31) çoğaltır.
    - {MONTH_NAME} ve {YEAR} etiketlerini günceller.
    - Orijinal TEMPLATE sayfasını silerek dosyayı kaydeder.
    """
    wb = openpyxl.load_workbook(config.MONTHLY_FLIGHT_TEMPLATE_PATH)
    
    if "TEMPLATE" not in wb.sheetnames:
        raise ValueError("Şablon dosyasında 'TEMPLATE' sayfası bulunamadı.")
    
    template_ws = wb["TEMPLATE"]
    month_tr = MONTHS_TR[month]
    
    # Ayın kaç gün çektiğini bul
    _, num_days = calendar.monthrange(year, month)
    
    for day in range(1, num_days + 1):
        # Örn: 01.07.2026
        sheet_name = f"{day:02d}.{month:02d}.{year}"
        
        # Sayfayı kopyala
        new_ws = wb.copy_worksheet(template_ws)
        import datetime
        new_ws.title = sheet_name
        
        # İlk satırdaki (G1 / R1C7) tarihi o güne eşitle
        # Orijinal şablonda 2026-06-01 vb. tarih objesi var.
        for row in range(1, 4):
            for col in range(1, 21):
                cell = new_ws.cell(row=row, column=col)
                if isinstance(cell.value, datetime.datetime):
                    cell.value = datetime.datetime(year, month, day)

    # Tüm günler oluşturulduktan sonra TEMPLATE'i sil
    del wb["TEMPLATE"]
    
    # Hedef konuma kaydet
    wb.save(save_path)
    wb.close()
