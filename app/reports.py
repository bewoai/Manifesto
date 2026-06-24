"""Driver route reports and operation output bundles."""
from __future__ import annotations

import re
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.manifest.planning import ReservationBlock, read_blocks


HEADERS = [
    "SIRA", "YOLCU ADI", "PAX", "OTEL", "ODA / İRTİBAT",
    "PICKUP", "ACENTE", "BALON", "GELECEĞİ YER",
]


def _safe_name(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", value.strip())
    return cleaned or "ATANMAMIS"


def driver_rows(blocks: list[ReservationBlock]) -> dict[str, list[dict]]:
    grouped_blocks: dict[str, list[ReservationBlock]] = defaultdict(list)
    for block in blocks:
        grouped_blocks[block.driver or "ATANMAMIŞ"].append(block)
    grouped: dict[str, list[dict]] = {}
    for driver, driver_blocks in grouped_blocks.items():
        rows: list[dict] = []
        driver_blocks.sort(key=lambda block: (
            block.pickup or "99:99", block.hotel, block.lead_row
        ))
        for block in driver_blocks:
            for index, passenger in enumerate(block.passengers):
                rows.append({
                    "name": passenger.name or "(kimlik eksik)",
                    "pax": block.pax if index == 0 else "",
                    "hotel": block.hotel,
                    "room": block.room,
                    "pickup": block.pickup,
                    "agency": block.agency,
                    "balloon": block.balloon,
                    "coming_place": block.coming_place,
                })
        grouped[driver] = rows
    return dict(sorted(grouped.items(), key=lambda item: item[0]))


def driver_summary(
    blocks: list[ReservationBlock],
    overrides: dict[str, str] | None = None,
) -> list[dict]:
    overrides = overrides or {}
    mapping: dict[str, set[str]] = defaultdict(set)
    for block in blocks:
        if block.balloon:
            mapping[block.balloon].add(block.driver or "ATANMAMIŞ")
    rows = []
    for balloon in sorted(mapping):
        default = ", ".join(sorted(mapping[balloon]))
        rows.append({
            "balloon": balloon,
            "driver": overrides.get(balloon, default),
        })
    return rows


def _excel_report(path: Path, sheet: str, driver: str, rows: list[dict]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YOL LİSTESİ"
    ws.merge_cells("A1:I1")
    ws["A1"] = f"İRTİFA - {sheet} - ŞOFÖR {driver}"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1D1430")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    thin = Side(style="thin", color="B9A9C8")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D97832")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for index, row in enumerate(rows, 1):
        values = [
            index, row["name"], row["pax"], row["hotel"], row["room"],
            row["pickup"], row["agency"], row["balloon"], row["coming_place"],
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(index + 3, col, value)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F4EEF8")

    widths = [7, 28, 8, 22, 18, 12, 22, 11, 20]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{max(3, len(rows) + 3)}"
    wb.save(path)
    wb.close()
    return path


def _register_pdf_fonts() -> tuple[str, str]:
    regular = Path(r"C:\Windows\Fonts\arial.ttf")
    bold = Path(r"C:\Windows\Fonts\arialbd.ttf")
    if regular.exists() and bold.exists():
        if "IrtifaArial" not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont("IrtifaArial", regular))
            pdfmetrics.registerFont(TTFont("IrtifaArial-Bold", bold))
        return "IrtifaArial", "IrtifaArial-Bold"
    return "Helvetica", "Helvetica-Bold"


def _pdf_report(path: Path, sheet: str, driver: str, rows: list[dict]) -> Path:
    regular, bold = _register_pdf_fonts()
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    title_style = ParagraphStyle(
        "title", fontName=bold, fontSize=15, alignment=TA_CENTER, spaceAfter=5 * mm
    )
    small = ParagraphStyle("cell", fontName=regular, fontSize=7.5, leading=9)
    data = [[Paragraph(header, small) for header in HEADERS]]
    for index, row in enumerate(rows, 1):
        values = [
            index, row["name"], row["pax"], row["hotel"], row["room"],
            row["pickup"], row["agency"], row["balloon"], row["coming_place"],
        ]
        data.append([Paragraph(str(value or ""), small) for value in values])
    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            10 * mm, 43 * mm, 12 * mm, 31 * mm, 29 * mm,
            22 * mm, 32 * mm, 17 * mm, 31 * mm,
        ],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D97832")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), bold),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#A895B8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4EEF8")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story = [
        Paragraph(f"İRTİFA - {sheet} - ŞOFÖR {driver}", title_style),
        Spacer(1, 2 * mm),
        table,
    ]
    document.build(story)
    return path


def export_driver_reports(
    planning_xlsx: Path,
    sheet: str,
    out_dir: Path,
    *,
    selected_driver: str = "",
) -> list[Path]:
    grouped = driver_rows(read_blocks(planning_xlsx, sheet))
    out_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[Path] = []
    for driver, rows in grouped.items():
        if selected_driver and driver != selected_driver:
            continue
        base = f"Sofor_{_safe_name(driver)}"
        outputs.append(_excel_report(out_dir / f"{base}.xlsx", sheet, driver, rows))
        outputs.append(_pdf_report(out_dir / f"{base}.pdf", sheet, driver, rows))
    return outputs


def export_summary(
    planning_xlsx: Path,
    sheet: str,
    out_dir: Path,
    *,
    overrides: dict[str, str] | None = None,
) -> list[Path]:
    rows = driver_summary(read_blocks(planning_xlsx, sheet), overrides)
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "Arac_Ozeti.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ARAÇ ÖZETİ"
    ws.append(["BALON", "ŞOFÖR / ARAÇ"])
    for row in rows:
        ws.append([row["balloon"], row["driver"]])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="D97832")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    wb.save(xlsx_path)
    wb.close()

    text_path = out_dir / "Arac_Ozeti.txt"
    text_path.write_text(
        "\n".join(f"{row['balloon']}: {row['driver']}" for row in rows),
        encoding="utf-8",
    )
    return [xlsx_path, text_path]


def zip_outputs(paths: list[Path], target: Path) -> Path:
    with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
        for path in paths:
            archive.write(path, arcname=path.name)
    return target
