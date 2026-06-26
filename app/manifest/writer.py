"""Manifesto türetme (brief §3): bir gün + bir balon kodu için planlama
satırlarını filtrele, 4 kolonu al, 2 dönüşüm uygula, şablona yaz.

Dönüşümler:
  AD SOYAD            <- col 4 (aynen, UPPERCASE)
  CİNSİYET           <- col 3 M/F  ->  M=Erkek, F=Kadın
  UYRUK              <- col 2 alpha-3  ->  ULKELER İngilizce adı (country_map)
  PASAPORT/KİMLİK NO <- col 20 (aynen)
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl
import sqlite3

from app import config
from app.country.country_map import iso3_to_name
from app.manifest.planning import PlanningRow, read_rows


@dataclass
class ManifestRow:
    name: str
    sex: str          # Erkek/Kadın
    nationality: str  # ULKELER İngilizce adı
    passport_no: str
    warnings: list[str]


def _transform(row: PlanningRow) -> ManifestRow:
    warnings: list[str] = []
    sex = config.SEX_TR.get(row.sex.upper().strip())
    if sex is None:
        warnings.append(f"bilinmeyen cinsiyet {row.sex!r}")
        sex = row.sex
    nat = iso3_to_name(row.nationality)
    if nat is None:
        warnings.append(f"uyruk kodu country_map'te yok: {row.nationality!r}")
        nat = row.nationality
    return ManifestRow(
        name=row.name.upper(),
        sex=sex,
        nationality=nat,
        passport_no=row.passport_no,
        warnings=warnings,
    )


def build_planning_rows_from_sqlite(flight_id: int, balloon: str, conn: sqlite3.Connection) -> list[PlanningRow]:
    """SQLite'tan belirli bir uçuş ve balonun yolcu satırlarını PlanningRow olarak döndürür."""
    balloon_upper = balloon.upper().strip()
    cursor = conn.execute(
        """
        SELECT 
            r.sort_order,
            r.pax,
            p.passport_no,
            p.nationality,
            p.sex,
            p.full_name,
            r.room_no,
            r.hotel,
            r.pickup_time,
            r.reserved_by,
            r.agency,
            r.flight_firm,
            r.balloon_code,
            r.pilot,
            r.notes,
            r.driver_no,
            r.destination,
            rp.seq
        FROM reservations r
        JOIN reservation_passengers rp ON rp.reservation_id = r.id
        JOIN passengers p ON rp.passenger_id = p.id
        WHERE r.flight_id = ? AND UPPER(r.balloon_code) = ?
        ORDER BY r.sort_order, rp.seq
        """,
        (flight_id, balloon_upper)
    )
    rows = []
    for row in cursor.fetchall():
        row_num = row["sort_order"] + row["seq"] - 1
        pax = row["pax"] if row["seq"] == 1 else None
        rows.append(PlanningRow(
            row=row_num,
            pax=pax,
            nationality=row["nationality"] or "",
            sex=row["sex"] or "",
            name=row["full_name"] or "",
            room=row["room_no"] or "",
            hotel=row["hotel"] or "",
            pickup=row["pickup_time"] or "",
            reserved_by=row["reserved_by"] or "",
            agency=row["agency"] or "",
            company=row["flight_firm"] or "",
            balloon=row["balloon_code"] or "",
            pilot=row["pilot"] or "",
            note=row["notes"] or "",
            driver=row["driver_no"] or "",
            coming_place=row["destination"] or "",
            passport_no=row["passport_no"] or "",
        ))
    return rows


def build_rows(planning_xlsx: Path, sheet: str, balloon: str) -> list[ManifestRow]:
    """Planlamadan tek balonun manifesto satırlarını üretir (yazmadan)."""
    balloon = balloon.upper().strip()
    
    from app.db import connect
    from app.manifest.importer import sheet_name_to_iso
    import sqlite3
    
    iso_date = sheet_name_to_iso(sheet)
    rows = None
    loaded_from_db = False
    
    if iso_date:
        try:
            conn = connect()
            flight = conn.execute("SELECT id FROM flights WHERE flight_date = ?", (iso_date,)).fetchone()
            if flight:
                res_count = conn.execute("SELECT COUNT(*) FROM reservations WHERE flight_id = ?", (flight["id"],)).fetchone()[0]
                if res_count > 0:
                    rows = build_planning_rows_from_sqlite(flight["id"], balloon, conn)
                    loaded_from_db = True
            conn.close()
        except Exception as e:
            print(f"Warning: Failed to load manifest rows from SQLite: {e}")
            
    if not loaded_from_db:
        rows = [r for r in read_rows(planning_xlsx, sheet) if r.balloon == balloon]
        
    return [_transform(r) for r in rows]



def write_manifest(rows: list[ManifestRow], out_path: Path,
                   template_path: Optional[Path] = None) -> Path:
    """ManifestRow'ları şablonun MANİFESTO sayfasına yazar (biçim + ULKELER korunur)."""
    template_path = template_path or config.MANIFEST_TEMPLATE_PATH
    wb = openpyxl.load_workbook(template_path)
    ws = wb[config.MANIFEST_SHEET]

    # mevcut veri satırlarını temizle (header'ın altından itibaren), biçimi bırak
    for r in range(config.MANIFEST_FIRST_DATA_ROW, ws.max_row + 1):
        for c in (config.M_COL_NAME, config.M_COL_SEX, config.M_COL_NATIONALITY, config.M_COL_PASSPORT):
            ws.cell(r, c).value = None

    for i, row in enumerate(rows):
        r = config.MANIFEST_FIRST_DATA_ROW + i
        ws.cell(r, config.M_COL_NAME).value = row.name
        ws.cell(r, config.M_COL_SEX).value = row.sex
        ws.cell(r, config.M_COL_NATIONALITY).value = row.nationality
        ws.cell(r, config.M_COL_PASSPORT).value = row.passport_no

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    wb.close()
    return out_path


def export(planning_xlsx: Path, sheet: str, balloon: str, out_dir: Path,
           template_path: Optional[Path] = None) -> Path:
    """Uçtan uca: filtrele -> dönüştür -> `<BALLOON>.xlsx` yaz. Çıktı yolunu döndürür."""
    rows = build_rows(planning_xlsx, sheet, balloon)
    out_path = Path(out_dir) / f"{balloon.upper().strip()}.xlsx"
    return write_manifest(rows, out_path, template_path)
