import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(1, 1, "Row 1")
ws.cell(2, 1, "Row 2")
ws.cell(3, 1, "Row 3")
ws.merge_cells("A2:B2")

print("Before:")
for r in range(1, 5):
    print(ws.cell(r, 1).value)

# Move Row 3 to Row 1
ws.insert_rows(1, 1)
# Now Row 3 is at Row 4
for c in range(1, 3):
    ws.cell(1, c).value = ws.cell(4, c).value
ws.delete_rows(4, 1)

print("After:")
for r in range(1, 5):
    print(ws.cell(r, 1).value)
