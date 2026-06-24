from pathlib import Path

import openpyxl

from app import config
from app.manifest.planning import create_reservation, write_identity, write_operation_details
from app.readiness import check_workbook
from app.reports import export_driver_reports


def _ready_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "24.06.2026"
    ws.cell(config.PLANNING_HEADER_ROW, config.COL_PAX).value = "PAX"
    ws.cell(config.PLANNING_HEADER_ROW, config.COL_NAME).value = "İSİM"
    wb.save(path)
    wb.close()
    result = create_reservation(
        path,
        "24.06.2026",
        pax=2,
        balloon="BYF",
        pilot="TAHSIN",
        fields={
            "hotel": "ARGOS",
            "pickup": "04:10",
            "agency": "ACME",
            "driver": "3",
            "coming_place": "DİREKT ALAN",
        },
    )
    write_identity(path, "24.06.2026", {
        result["rows"][0]: {
            "nationality": "TUR", "sex": "M", "name": "ALİ VELİ", "passport_no": "P1",
        },
        result["rows"][1]: {
            "nationality": "TUR", "sex": "F", "name": "AYŞE VELİ", "passport_no": "P2",
        },
    })


def test_ready_check_and_driver_outputs(tmp_path):
    path = tmp_path / "planning.xlsx"
    _ready_workbook(path)
    readiness = check_workbook(path, "24.06.2026", balloon_capacity=28)
    assert readiness["ready"]
    outputs = export_driver_reports(path, "24.06.2026", tmp_path / "out")
    assert {item.suffix for item in outputs} == {".xlsx", ".pdf"}
    assert all(item.exists() and item.stat().st_size > 0 for item in outputs)
