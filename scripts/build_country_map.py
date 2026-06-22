"""ULKELER sayfasından country_map.json üretir (brief §14.2).

Her `ULKELER` İngilizce adını ISO 3166-1 alpha-3 koduna çözer:
  1) elle override tablosu (yazım hataları / özel adlar / ISO-dışı kodlar),
  2) pycountry tam ad / common_name / official_name,
  3) pycountry fuzzy arama (son çare).

Kullanım:
  python -m scripts.build_country_map --source "<manifest_or_template.xlsx>"
Kaynak verilmezse SOURCE_MANIFEST_XLSX (config) kullanılır.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
import pycountry

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app.config import COUNTRY_MAP_PATH, SOURCE_MANIFEST_XLSX  # noqa: E402

# `ULKELER`'deki tam string -> ISO alpha-3 (yazım hataları ve özel durumlar dâhil).
OVERRIDES: dict[str, str] = {
    "Bolivia": "BOL",
    "Bonaire, Sint Eustatius and Saba": "BES",
    "Brunei": "BRN",
    "Cape Verde": "CPV",
    "Congo": "COG",
    "Democratic Republic of the Congo": "COD",
    "Curacao": "CUW",
    "Czech Republic": "CZE",
    "Ivory Coast": "CIV",
    "Falkland Islands (Malvinas)": "FLK",
    "Guadaloupe": "GLP",        # yazım: Guadeloupe
    "Heard Island and McDonald Islands": "HMD",
    "Iran": "IRN",
    "Kosovo": "XKX",           # ISO-dışı, kullanıcı atamalı
    "Laos": "LAO",
    "Macao": "MAC",
    "Macedonia": "MKD",
    "Micronesia": "FSM",
    "Moldava": "MDA",          # yazım: Moldova
    "Myanmar (Burma)": "MMR",
    "North Korea": "PRK",
    "Palestine": "PSE",
    "Phillipines": "PHL",      # yazım: Philippines
    "Russia": "RUS",
    "Saint Barthelemy": "BLM",
    "Saint Helena": "SHN",
    "Saint Martin": "MAF",
    "Saint Pierre and Miquelon": "SPM",
    "Sint Maarten": "SXM",
    "South Korea": "KOR",
    "Swaziland": "SWZ",
    "Syria": "SYR",
    "Taiwan": "TWN",
    "Tanzania": "TZA",
    "Timor-Leste (East Timor)": "TLS",
    "Türkiye": "TUR",
    "Vatican City": "VAT",
    "Venezuela": "VEN",
    "Vietnam": "VNM",
    "Virgin Islands, British": "VGB",
    "Virgin Islands, US": "VIR",
}

# `ULKELER`'de eksik olduğunu bildiğimiz ama yine de haritaya eklenmesi gereken kayıtlar.
# (Kaynak sayfada Iceland ile Indonesia arasında "India" satırı boş bırakılmış.)
EXTRA: dict[str, str] = {
    "IND": "India",
}


def resolve_alpha3(name: str) -> str | None:
    if name in OVERRIDES:
        return OVERRIDES[name]
    # pycountry tam alan eşleşmeleri
    for getter in ({"name": name}, {"common_name": name}, {"official_name": name}):
        try:
            rec = pycountry.countries.get(**getter)
        except Exception:
            rec = None
        if rec:
            return rec.alpha_3
    # son çare: fuzzy
    try:
        return pycountry.countries.search_fuzzy(name)[0].alpha_3
    except LookupError:
        return None


def read_ulkeler(source: Path) -> list[str]:
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    if "ULKELER" not in wb.sheetnames:
        raise SystemExit(f"'{source}' içinde ULKELER sayfası yok")
    ws = wb["ULKELER"]
    names = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip():
            names.append(str(v).strip())
        else:
            print(f"  [uyarı] ULKELER r{r}: boş satır (kaynak veride eksik kayıt)")
    wb.close()
    return names


def build(source: Path) -> dict[str, dict]:
    names = read_ulkeler(source)
    mapping: dict[str, dict] = {}
    unresolved: list[str] = []
    collisions: list[tuple[str, str, str]] = []

    for name in names:
        code = resolve_alpha3(name)
        if not code:
            unresolved.append(name)
            continue
        if code in mapping and mapping[code]["name"] != name:
            collisions.append((code, mapping[code]["name"], name))
        mapping[code] = {"code": code, "name": name}

    for code, name in EXTRA.items():
        mapping.setdefault(code, {"code": code, "name": name})

    print(f"\nÇözülen: {len(mapping)} kod | ULKELER ad sayısı: {len(names)}")
    if unresolved:
        print("ÇÖZÜLEMEYEN adlar:", unresolved)
    if collisions:
        print("ÇAKIŞMALAR (aynı alpha-3, farklı ad):", collisions)
    return dict(sorted(mapping.items()))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", type=Path, default=SOURCE_MANIFEST_XLSX,
                    help="ULKELER sayfasını içeren manifest/şablon xlsx")
    ap.add_argument("--out", type=Path, default=COUNTRY_MAP_PATH)
    args = ap.parse_args()

    mapping = build(args.source)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2, sort_keys=True)
    print(f"Yazıldı -> {args.out} ({len(mapping)} kayıt)")


if __name__ == "__main__":
    main()
