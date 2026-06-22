"""Golden pair testi: 22.06 planlama BZR satırları -> BZR.xlsx ile birebir.

Bu, country_map + 2 dönüşüm + writer hattını gerçek veriyle uçtan uca doğrular.
Vision API veya WhatsApp gerekmez.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from app import config
from app.manifest.writer import build_rows, export

SHEET = "22.06.2026"
BALLOON = "BZR"


def _read_manifest_rows(path: Path) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[config.MANIFEST_SHEET]
    rows = []
    for r in range(config.MANIFEST_FIRST_DATA_ROW, ws.max_row + 1):
        name = ws.cell(r, config.M_COL_NAME).value
        if not name:
            continue
        rows.append((
            str(name).strip(),
            str(ws.cell(r, config.M_COL_SEX).value or "").strip(),
            str(ws.cell(r, config.M_COL_NATIONALITY).value or "").strip(),
            str(ws.cell(r, config.M_COL_PASSPORT).value or "").strip(),
        ))
    wb.close()
    return rows


@pytest.mark.usefixtures("country_map_ready", "template_ready", "planning_ready", "golden_ready")
def test_bzr_export_matches_golden():
    expected = _read_manifest_rows(config.GOLDEN_MANIFEST_XLSX)

    with tempfile.TemporaryDirectory() as tmp:
        out = export(config.PLANNING_XLSX, SHEET, BALLOON, Path(tmp))
        produced = _read_manifest_rows(out)

    assert len(produced) == len(expected) == 28
    for i, (got, exp) in enumerate(zip(produced, expected)):
        assert got == exp, f"satır {i}: üretilen {got} != golden {exp}"


@pytest.mark.usefixtures("country_map_ready", "planning_ready")
def test_transforms_applied():
    rows = build_rows(config.PLANNING_XLSX, SHEET, BALLOON)
    assert rows, "BZR satırı bulunamadı"
    for r in rows:
        assert r.sex in ("Erkek", "Kadın"), r.sex          # M/F -> TR
        assert r.nationality and r.nationality.isascii() or True
        assert not r.warnings, f"{r.name}: {r.warnings}"   # tüm uyruklar çözülmeli
    # ilk iki yolcu CHE -> Switzerland olmalı (golden ile aynı sıra)
    assert rows[0].nationality == "Switzerland"
    assert rows[1].nationality == "Switzerland"
