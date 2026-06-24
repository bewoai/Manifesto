from pathlib import Path

import openpyxl
import pytest

from app import config
from app.excel_store import WorkbookConflictError, atomic_update, workbook_revision
from app.manifest.planning import (
    create_reservation,
    delete_passenger_from_reservation,
    read_blocks,
    resize_reservation,
    write_identity,
)


def _planning(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "24.06.2026"
    ws.cell(config.PLANNING_HEADER_ROW, config.COL_PAX).value = "PAX"
    ws.cell(config.PLANNING_HEADER_ROW, config.COL_NAME).value = "İSİM"
    wb.save(path)
    wb.close()


def test_atomic_update_rejects_stale_revision_and_creates_backup(tmp_path):
    path = tmp_path / "planning.xlsx"
    _planning(path)
    revision = workbook_revision(path)

    result, new_revision, backup = atomic_update(
        path,
        expected_revision=revision,
        reason="test",
        mutator=lambda temp: create_reservation(
            temp, "24.06.2026", pax=2, balloon="BYF"
        ),
    )
    assert result["rows"] == [4, 5]
    assert new_revision != revision
    assert backup.exists()

    with pytest.raises(WorkbookConflictError):
        atomic_update(
            path,
            expected_revision=revision,
            reason="stale",
            mutator=lambda temp: temp,
        )


def test_delete_middle_passenger_compacts_and_next_reservation_is_safe(tmp_path):
    path = tmp_path / "planning.xlsx"
    _planning(path)
    result = create_reservation(path, "24.06.2026", pax=3, balloon="BYF")
    write_identity(path, "24.06.2026", {
        4: {"name": "ONE", "passport_no": "P1"},
        5: {"name": "TWO", "passport_no": "P2"},
        6: {"name": "THREE", "passport_no": "P3"},
    })

    delete_passenger_from_reservation(
        path, "24.06.2026", lead_row=4, rows=result["rows"], target_row=5
    )
    block = read_blocks(path, "24.06.2026")[0]
    assert [p.name for p in block.passengers] == ["ONE", "THREE"]
    created = create_reservation(path, "24.06.2026", pax=1, balloon="BTK")
    assert created["lead_row"] == 6


def test_resize_requires_confirmation_for_identity_loss(tmp_path):
    path = tmp_path / "planning.xlsx"
    _planning(path)
    result = create_reservation(path, "24.06.2026", pax=2, balloon="BYF")
    write_identity(path, "24.06.2026", {5: {"name": "SECOND", "passport_no": "P2"}})
    with pytest.raises(ValueError, match="Açık onay"):
        resize_reservation(
            path, "24.06.2026", lead_row=4, rows=result["rows"], new_pax=1
        )
    rows = resize_reservation(
        path,
        "24.06.2026",
        lead_row=4,
        rows=result["rows"],
        new_pax=1,
        allow_data_loss=True,
    )
    assert rows == [4]
