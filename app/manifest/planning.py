"""Planlama xlsx okuma + (hibrit mimari) sadece 4 kimlik kolonunu geri yazma.

Brief §12.1 kararı: Excel kaynak olarak KALIR. Sistem planlama sayfasını okur,
yalnızca col 2/3/4/20'yi (UYRUK, M/F, İSİM, PASAPORT NO) doldurup geri yazar;
diğer her şey (PAX, otel, balon, pilot...) operasyon girişidir ve dokunulmaz.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

from app import config


@dataclass
class PlanningRow:
    row: int
    pax: Optional[int]           # rezervasyon grubu boyutu (yalnız lead satırda)
    nationality: str             # col 2
    sex: str                     # col 3
    name: str                    # col 4
    room: str
    hotel: str                   # col 6
    pickup: str                  # col 7
    reserved_by: str
    agency: str                  # col 9 — eşleştirme anahtarı
    company: str
    balloon: str                 # col 11 — manifesto filtresi
    pilot: str
    note: str
    driver: str
    coming_place: str
    passport_no: str             # col 20


@dataclass
class PassengerIdentity:
    row: int
    nationality: str
    sex: str
    name: str
    passport_no: str


@dataclass
class ReservationBlock:
    """Bir rezervasyon: lead satır (PAX dolu) + altındaki boş-PAX yolcu satırları.

    Operatör bu bloğu seçer, ait olduğu pasaport fotoğraflarını yükler; onaylanan
    pasaportlar bu bloğun satırlarına (col 2/3/4/20) sırayla yazılır.
    """
    lead_row: int
    rows: list[int]              # bloğun tüm satır indexleri (lead dâhil)
    pax: Optional[int]           # beklenen yolcu sayısı (PAX checksum)
    room: str
    agency: str
    reserved_by: str
    company: str
    balloon: str
    pilot: str
    note: str
    driver: str
    coming_place: str
    hotel: str
    pickup: str
    lead_name: str               # mevcut booking ismi (varsa) — eşleştirme ipucu
    passengers: list[PassengerIdentity] = field(default_factory=list)

    def label(self) -> str:
        pax = self.pax if self.pax is not None else "?"
        who = self.lead_name or "(isim yok)"
        return (f"r{self.lead_row}  •  {self.agency or '—'}  •  {self.balloon or '—'}  "
                f"•  PAX {pax}  •  {self.hotel or '—'}  •  {who}")


def _cell(ws, row: int, col: int) -> str:
    v = ws.cell(row, col).value
    return "" if v is None else str(v).strip()


def read_rows(planning_xlsx: Path, sheet: str) -> list[PlanningRow]:
    """Bir gün sayfasının tüm yolcu satırlarını okur (header row 3 sonrası)."""
    wb = openpyxl.load_workbook(planning_xlsx, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Planlama dosyasında '{sheet}' sayfası yok. Mevcut: {wb.sheetnames}")
    ws = wb[sheet]
    rows: list[PlanningRow] = []
    for r in range(config.PLANNING_FIRST_DATA_ROW, ws.max_row + 1):
        name = _cell(ws, r, config.COL_NAME)
        balloon = _cell(ws, r, config.COL_BALLOON)
        # tamamen boş satırları atla (isim ve balon yoksa veri yok say)
        if not name and not balloon and not _cell(ws, r, config.COL_PASSPORT_NO):
            continue
        pax_raw = _cell(ws, r, config.COL_PAX)
        try:
            pax = int(float(pax_raw)) if pax_raw else None
        except ValueError:
            pax = None
        rows.append(PlanningRow(
            row=r,
            pax=pax,
            nationality=_cell(ws, r, config.COL_UYRUK),
            sex=_cell(ws, r, config.COL_MF),
            name=name,
            room=_cell(ws, r, config.COL_ROOM),
            hotel=_cell(ws, r, config.COL_HOTEL),
            pickup=_cell(ws, r, config.COL_PICKUP),
            reserved_by=_cell(ws, r, config.COL_RESERVED_BY),
            agency=_cell(ws, r, config.COL_AGENCY),
            company=_cell(ws, r, config.COL_COMPANY),
            balloon=balloon.upper(),
            pilot=_cell(ws, r, config.COL_PILOT),
            note=_cell(ws, r, config.COL_NOTE),
            driver=_cell(ws, r, config.COL_DRIVER),
            coming_place=_cell(ws, r, config.COL_COMING_PLACE),
            passport_no=_cell(ws, r, config.COL_PASSPORT_NO),
        ))
    wb.close()
    return rows


def group_blocks(rows: list[PlanningRow]) -> list[ReservationBlock]:
    """Satırları rezervasyon bloklarına böler: PAX dolu satır yeni blok başlatır."""
    blocks: list[ReservationBlock] = []
    cur: Optional[ReservationBlock] = None
    for pr in rows:
        starts = pr.pax is not None
        if starts or cur is None:
            cur = ReservationBlock(
                lead_row=pr.row, rows=[pr.row], pax=pr.pax,
                room=pr.room, agency=pr.agency, reserved_by=pr.reserved_by,
                company=pr.company, balloon=pr.balloon, pilot=pr.pilot,
                note=pr.note, driver=pr.driver, coming_place=pr.coming_place,
                hotel=pr.hotel,
                pickup=pr.pickup, lead_name=pr.name,
                passengers=[PassengerIdentity(
                    row=pr.row,
                    nationality=pr.nationality,
                    sex=pr.sex,
                    name=pr.name,
                    passport_no=pr.passport_no,
                )],
            )
            blocks.append(cur)
        else:
            cur.rows.append(pr.row)
            cur.passengers.append(PassengerIdentity(
                row=pr.row,
                nationality=pr.nationality,
                sex=pr.sex,
                name=pr.name,
                passport_no=pr.passport_no,
            ))
    return blocks


def read_blocks(planning_xlsx: Path, sheet: str) -> list[ReservationBlock]:
    return group_blocks(read_rows(planning_xlsx, sheet))


# ─────────────────────────────────────────────────────────────────────
#  Balon otomatik atama (brief §15 revize: sistem atar)
# ─────────────────────────────────────────────────────────────────────

def block_seats(block: ReservationBlock) -> int:
    """Bloğun kapladığı koltuk sayısı (PAX, yoksa satır sayısı)."""
    return block.pax if block.pax is not None else len(block.rows)


def balloon_load(blocks: list[ReservationBlock]) -> dict[str, int]:
    """Balon kodu → dolu koltuk sayısı."""
    load: dict[str, int] = {}
    for b in blocks:
        if b.balloon:
            load[b.balloon] = load.get(b.balloon, 0) + block_seats(b)
    return load


def assign_balloon(
    load: dict[str, int],
    pax: int,
    codes: list[str],
    capacity: int,
) -> tuple[str, bool]:
    """Yeni bir `pax` kişilik grup için balon seç (grup bölünmez).

    First-fit: `codes` sırasında kalan kapasitesi ≥ pax olan ilk balonu döndürür.
    Hiçbiri sığmazsa en az dolu balonu döndürür ve overflow=True işaretler.
    """
    if not codes:
        return "", True
    for code in codes:
        if capacity - load.get(code, 0) >= pax:
            return code, False
    # Hiçbiri tüm grubu alamıyor → en boş balon + overflow uyarısı
    least = min(codes, key=lambda c: load.get(c, 0))
    return least, True


def pilot_for_balloon(rows: list[PlanningRow], code: str) -> str:
    """Verilen balonun mevcut satırlarından pilot adını bul (yoksa boş)."""
    code = code.upper().strip()
    for pr in rows:
        if pr.balloon == code and pr.pilot:
            return pr.pilot
    return ""


def next_free_row(ws) -> int:
    """Ad/balon/pasaport hepsi boş ilk veri satırı (yoksa max_row+1)."""
    for r in range(config.PLANNING_FIRST_DATA_ROW, ws.max_row + 1):
        name = _cell(ws, r, config.COL_NAME)
        balloon = _cell(ws, r, config.COL_BALLOON)
        ppno = _cell(ws, r, config.COL_PASSPORT_NO)
        if not name and not balloon and not ppno:
            return r
    return ws.max_row + 1


def create_reservation(
    planning_xlsx: Path,
    sheet: str,
    *,
    pax: int,
    balloon: str,
    pilot: str = "",
    fields: Optional[dict] = None,
    out_path: Optional[Path] = None,
) -> dict:
    """Günün sayfasına boş satırlardan itibaren yeni rezervasyon bloğu ekler.

    Lider satıra PAX + lider-özel alanlar (agency/hotel/pickup/room/reserved_by/note);
    bloğun tüm satırlarına balloon/pilot/driver/coming_place yazılır. Biçim korunur.
    """
    fields = dict(fields or {})
    fields["balloon"] = balloon
    if pilot:
        fields["pilot"] = pilot

    wb = openpyxl.load_workbook(planning_xlsx)
    ws = wb[sheet]
    lead = next_free_row(ws)
    rows = list(range(lead, lead + max(1, int(pax))))
    last = rows[-1]

    # Hedef satırlarla kesişen mevcut birleşmeleri çöz (kaynak günden miras kalan,
    # yeni blok sınırlarımıza uymayan birleşik hücreler yazımı engeller).
    for rng in list(ws.merged_cells.ranges):
        if not (rng.max_row < lead or rng.min_row > last):
            ws.unmerge_cells(str(rng))

    ws.cell(lead, config.COL_PAX).value = int(pax)
    for key, col in config.OPERATION_FIELD_TO_COL.items():
        if key == "pax" or key not in fields:
            continue
        value = fields[key]
        if value in (None, ""):
            continue
        target_rows = rows if key in config.BLOCK_WIDE_OPERATION_FIELDS else [lead]
        for row in target_rows:
            ws.cell(row, col).value = value

    # Lider-gösterim kolonlarını blok boyunca dikey birleştir (orijinal görünüm)
    if len(rows) > 1:
        for col in config.LEAD_MERGE_COLS:
            ws.merge_cells(start_row=lead, start_column=col, end_row=last, end_column=col)

    target = out_path or planning_xlsx
    wb.save(target)
    wb.close()
    return {"lead_row": lead, "rows": rows, "balloon": balloon, "pilot": pilot}


def list_sheets(planning_xlsx: Path) -> list[str]:
    """Gün sayfalarını döndürür (Sayfa2 gibi özet sayfaları hariç)."""
    wb = openpyxl.load_workbook(planning_xlsx, read_only=True)
    names = [s for s in wb.sheetnames if s.lower() not in ("sayfa2", "sheet2")]
    wb.close()
    return names


def create_day_sheet(planning_xlsx: Path, new_sheet: str, source_sheet: Optional[str] = None) -> Path:
    """Yeni gün sekmesi oluşturur: kaynak günü kopyalar, veri satırlarını temizler."""
    wb = openpyxl.load_workbook(planning_xlsx)
    if new_sheet in wb.sheetnames:
        wb.close()
        raise ValueError(f"'{new_sheet}' sayfası zaten var.")

    candidates = [s for s in wb.sheetnames if s.lower() not in ("sayfa2", "sheet2")]
    if source_sheet and source_sheet in wb.sheetnames:
        source_name = source_sheet
    elif candidates:
        source_name = candidates[-1]
    else:
        wb.close()
        raise ValueError("Kopyalanacak gün sayfası bulunamadı.")

    ws = wb.copy_worksheet(wb[source_name])
    ws.title = new_sheet
    for row in range(config.PLANNING_FIRST_DATA_ROW, ws.max_row + 1):
        for col in range(1, config.COL_PASSPORT_NO + 1):
            ws.cell(row, col).value = None

    wb.save(planning_xlsx)
    wb.close()
    return planning_xlsx


def delete_reservation(planning_xlsx: Path, sheet: str, rows: list[int],
                       out_path: Optional[Path] = None) -> Path:
    """Bir rezervasyon bloğunun satırlarını temizler (tüm kolonlar), birleşmeleri çözer.

    Satırlar boşaldığı için sonraki `create_reservation` bunları tekrar kullanabilir.
    """
    if not rows:
        return planning_xlsx
    wb = openpyxl.load_workbook(planning_xlsx)
    ws = wb[sheet]
    lo, hi = min(rows), max(rows)
    # Bu satırlarla kesişen birleşmeleri çöz (yoksa MergedCell yazılamaz)
    for rng in list(ws.merged_cells.ranges):
        if not (rng.max_row < lo or rng.min_row > hi):
            ws.unmerge_cells(str(rng))
    for r in rows:
        for c in range(1, config.COL_PASSPORT_NO + 1):
            ws.cell(r, c).value = None
    target = out_path or planning_xlsx
    wb.save(target)
    wb.close()
    return target


def delete_passenger_from_reservation(
    planning_xlsx: Path,
    sheet: str,
    *,
    lead_row: int,
    rows: list[int],
    target_row: int,
    out_path: Optional[Path] = None,
) -> dict:
    """Rezervasyon içinden tek yolcu siler, operasyon alanlarını korur.

    Lider yolcu silinirse alttaki ilk dolu kimlik lider satıra taşınır. Böylece
    PAX/otel/acente gibi lider satırdaki operasyon bilgileri kaybolmaz.
    """
    if target_row not in rows:
        raise ValueError("Silinecek yolcu bu rezervasyon bloğunda değil.")
    if len(rows) <= 1:
        delete_reservation(planning_xlsx, sheet, rows, out_path=out_path)
        return {"deleted_block": True, "pax": 0}

    wb = openpyxl.load_workbook(planning_xlsx)
    ws = wb[sheet]
    lo, hi = min(rows), max(rows)

    for rng in list(ws.merged_cells.ranges):
        if not (rng.max_row < lo or rng.min_row > hi):
            ws.unmerge_cells(str(rng))

    identity_cols = (config.COL_UYRUK, config.COL_MF, config.COL_NAME, config.COL_PASSPORT_NO)

    def has_identity(row: int) -> bool:
        return any(_cell(ws, row, col) for col in identity_cols)

    clear_row = target_row
    if target_row == lead_row:
        replacement = next((r for r in rows if r != lead_row and has_identity(r)), None)
        if replacement is not None:
            for col in identity_cols:
                ws.cell(lead_row, col).value = ws.cell(replacement, col).value
            clear_row = replacement
        else:
            for col in identity_cols:
                ws.cell(lead_row, col).value = None
            clear_row = None

    if clear_row is not None:
        for col in range(1, config.COL_PASSPORT_NO + 1):
            ws.cell(clear_row, col).value = None

    pax_raw = _cell(ws, lead_row, config.COL_PAX)
    try:
        pax = int(float(pax_raw)) if pax_raw else len(rows)
    except ValueError:
        pax = len(rows)
    pax = max(1, pax - 1)
    ws.cell(lead_row, config.COL_PAX).value = pax

    target = out_path or planning_xlsx
    wb.save(target)
    wb.close()
    return {"deleted_block": False, "pax": pax, "cleared_row": clear_row}


def write_identity(planning_xlsx: Path, sheet: str, updates: dict[int, dict],
                   out_path: Optional[Path] = None) -> Path:
    """Belirli satırların yalnız kimlik kolonlarını günceller, biçimi korur.

    updates: { row_index: {"nationality": "...", "sex": "M",
                           "name": "...", "passport_no": "..."} }
    Yalnızca verilen alanlar yazılır; None/eksik alanlara dokunulmaz.
    """
    wb = openpyxl.load_workbook(planning_xlsx)  # biçim korunur
    ws = wb[sheet]
    field_to_col = {
        "nationality": config.COL_UYRUK,
        "sex": config.COL_MF,
        "name": config.COL_NAME,
        "passport_no": config.COL_PASSPORT_NO,
    }
    for row, fields in updates.items():
        for key, col in field_to_col.items():
            if key in fields and fields[key] is not None:
                ws.cell(row, col).value = fields[key]
    target = out_path or planning_xlsx
    wb.save(target)
    wb.close()
    return target


def write_operation_details(
    planning_xlsx: Path,
    sheet: str,
    lead_row: int,
    rows: list[int],
    fields: dict,
    out_path: Optional[Path] = None,
) -> Path:
    """Rezervasyonun operasyon alanlarını yazar.

    PAX/otel/acente gibi rezervasyon alanları lead satıra yazılır. Balon, pilot,
    şoför ve geleceği yer bloktaki tüm yolculara yazılır; bu kolonlar gerçek
    listede her yolcu satırında tekrar ediyor ve manifesto filtresi için önemlidir.
    """
    wb = openpyxl.load_workbook(planning_xlsx)
    ws = wb[sheet]
    for key, col in config.OPERATION_FIELD_TO_COL.items():
        if key not in fields:
            continue
        value = fields[key]
        target_rows = rows if key in config.BLOCK_WIDE_OPERATION_FIELDS else [lead_row]
        for row in target_rows:
            ws.cell(row, col).value = value
    target = out_path or planning_xlsx
    wb.save(target)
    wb.close()
    return target
